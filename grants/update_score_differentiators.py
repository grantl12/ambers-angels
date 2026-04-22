import openpyxl

file_path = 'D:/Users/grant/Documents/AmbersAngels/grants/SCORE Financial Projections Template.xlsx'
wb = openpyxl.load_workbook(file_path)

ws5 = wb['5a-OpExYear1']
# Column O is column 15
ws5.cell(row=8, column=15).value = "STRATEGIC DIFFERENTIATORS"
ws5.cell(row=9, column=15).value = "Market Advantage: Filling gaps in fixed ALPR with mobile nodes."
ws5.cell(row=10, column=15).value = "Sustainability: Low overhead via community-volunteered hardware."

wb.save(file_path)
print("Excel template updated with differentiator notes.")
