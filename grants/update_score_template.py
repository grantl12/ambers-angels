import openpyxl

file_path = 'D:/Users/grant/Documents/AmbersAngels/grants/SCORE Financial Projections Template.xlsx'
wb = openpyxl.load_workbook(file_path)

# Update Starting Point
# 1-StartingPoint sheet
ws1 = wb['1-StartingPoint']
# The dataframe read showed Row 3 (0-indexed) has Company Name and Prepared By
# Row 4 in Excel = index 3
ws1['C4'] = "Amber's Angels"
ws1['B4'] = "Grant"

# Fixed Assets - Equipment is index 10 in DataFrame, so Row 11 in Excel
ws1['C11'] = 2500  # Equipment (Mavic 3, etc)

# Operating Capital
# Legal is index 20 in DataFrame? Wait let's check the previous read.
# 17: Pre-Opening Salaries
# 18: Prepaid Insurance
# 19: Inventory
# 20: Legal and Other Professional Fees
ws1['C21'] = 400   # IRS, SOS Fees

# Cash on Hand - let's use C26 for starting infrastructure/server fund
ws1['C26'] = 1200

# Update OpEx Year 1
# 5a-OpExYear1 sheet
ws5 = wb['5a-OpExYear1']
# In the previous read:
# Row 12 (index 11): Insurance
# Row 13 (index 12): Legal and Professional
# Row 14 (index 13): Licenses
# Row 15 (index 14): Office Expense
# Row 16 (index 15): Rent or Lease -- Vehicles, Machinery, Equipment
# Row 17 (index 16): Rent or Lease -- Other Business Property (Servers?)
# Row 18 (index 17): Repairs

# Months 1-12 are Columns C to N (3 to 14)
for col in range(3, 15):
    ws5.cell(row=13, column=col).value = 66.67 # Insurance ~$800/yr
    ws5.cell(row=17, column=col).value = 100   # Infrastructure ~$1200/yr

# Professional Services (In-Kind)
# We can use 'Contract Labor' (Row 12) for the valuation, 
# but usually, SCORE templates are for cash flow.
# I will put a note in the 'Miscellaneous' (Row 23)
ws5.cell(row=23, column=3).value = "In-Kind Prof Services: $81,000/yr (SDVOSB)"

wb.save(file_path)
print("Excel template updated successfully.")
